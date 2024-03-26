##O progama scanneia a lista de RG's captura a data de expedição, e identifica as pessoas que estão com o RG fora da validade
##Isto é, quem possui idade de 0 a 12 anos, o documento tem validade de 5 anos. Já de 12 a 60 anos incompletos, o prazo é de 10 anos e a partir dos 60 anos, o documento passa a ter validade indeterminada.

from datetime import date
import pytesseract
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import cv2
import os

cont = 2


workbook = load_workbook('DADOS.xlsx')
planilha = workbook['Plan1']
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

#caminho do tesseract
caminho = r"C:\Program Files\Tesseract-OCR" 


#definindo a variavel de ambiente direto pelo codigo
os.environ['TESSDATA_PREFIX'] = r'C:\Program Files\Tesseract-OCR\tessdata' 


#executavel na pasta do tesseract
pytesseract.pytesseract.tesseract_cmd = caminho + r'\tesseract.exe' 


#pasta com os documentos
pasta_rgs = os.listdir(r"D:\Nova pasta\projetos\planilhamento de documento de identidade\rgs") 
def calcularValidade(data_expedicao,idade):
    hoje = date.today()
    
    ano_atual = hoje.year
    mes_atual = hoje.month
    dia_atual = hoje.day
    
    dia_expedicao, mes_expedicao, ano_expedicao = data_expedicao.split('/')
    dia_expedicao = int(dia_expedicao)
    mes_expedicao = int(mes_expedicao)
    ano_expedicao = int(ano_expedicao)
    
    tempo_de_rg = ano_atual-ano_expedicao

    if idade <12 and tempo_de_rg >5:
        return 0
    
    elif idade > 1 and idade < 60 and tempo_de_rg > 10:
        return 0
    
    else:
        return 1


    
def calcularIdade(data_nascimento):
    hoje = date.today()
    ano_atual = hoje.year
    mes_atual = hoje.month
    dia_atual = hoje.day

    dia_nascimento, mes_nascimento, ano_nascimento = data_nascimento.split('/')
    dia_nascimento = int(dia_nascimento)
    mes_nascimento = int(mes_nascimento)
    ano_nascimento = int(ano_nascimento)

    idade = ano_atual - (ano_nascimento)

    # Verificar se o aniversário já ocorreu neste ano
    if (mes_atual, dia_atual) < (mes_nascimento, dia_nascimento):
        idade -= 1

    return idade


##capturando as informações do RG e transformando em string
for rg in pasta_rgs:
    imagem_rg = cv2.imread(f"rgs/{rg}")
    posicao_nome = imagem_rg[918:918+80, 162:162+300]
    string_nome = pytesseract.image_to_string(posicao_nome,lang='por')
    
    posicao_data_nasc = imagem_rg[1197:1197+80, 800:800+300]
    string_data_nasc = pytesseract.image_to_string(posicao_data_nasc)
    
    posicao_data_exped = imagem_rg[850:850+80, 788:788+300]
    string_data_exped = pytesseract.image_to_string(posicao_data_exped)
    
        
    posicao_registro = imagem_rg[850:850+80, 210:210+300]
    string_registro = pytesseract.image_to_string(posicao_registro)
    
    posicao_cpf = imagem_rg[1360:1360+80, 150:150+300]
    string_cpf = pytesseract.image_to_string(posicao_cpf)
    
    print(f"NOME: {string_nome}")
    print(f"DATA DE NASCIMENTO: {string_data_nasc}")
    print(f"DATA DE EXPEDIÇÃO: {string_data_exped}")
    print(f"REGISTRO: {string_registro}")
    print(f"CPF: {string_cpf}")
    
    idade = calcularIdade(string_data_nasc)
    planilha[f'A{cont}'] = string_nome
    planilha[f'C{cont}'] = string_data_nasc
    planilha[f'D{cont}'] = string_registro
    planilha[f'E{cont}'] = string_data_exped
    planilha[f'B{cont}'] = idade
    if calcularValidade(string_data_exped,idade):
        planilha[f'F{cont}'].fill = green_fill
        
    else:
        planilha[f'F{cont}'].fill = red_fill
        
        
    cont = cont + 1 
    
    
    workbook.save('DADOS.xlsx')